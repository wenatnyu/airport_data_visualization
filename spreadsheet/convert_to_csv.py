import openpyxl
import pandas as pd
from datetime import datetime, timedelta
import re
import os

def evaluate_formula(value):
    try:
        # 如果不是字符串或不包含运算符，直接返回原值
        if not isinstance(value, str) or not any(op in value for op in ['+', '-', '*', '/']):
            return value
            
        # 替换中文数字
        chinese_nums = {'一': '1', '二': '2', '三': '3', '四': '4', '五': '5',
                       '六': '6', '七': '7', '八': '8', '九': '9', '零': '0'}
        for cn, num in chinese_nums.items():
            value = value.replace(cn, num)
        
        # 移除除数字和运算符外的所有字符
        value = re.sub(r'[^0-9+\-*/().]', '', value)
        
        # 计算公式
        return eval(value)
    except:
        return value

def process_date(value, year):
    try:
        # 处理Excel数字日期格式
        if isinstance(value, (int, float)):
            # Excel的日期是从1900年1月1日开始的天数
            excel_date = datetime(1900, 1, 1) + timedelta(days=int(value)-2)  # -2 是因为Excel的日期系统有一个bug
            # 返回完整的年月日格式
            return excel_date.strftime('%Y-%m-%d')
        # 处理字符串日期格式
        elif isinstance(value, str):
            # 使用正则表达式提取月和日
            match = re.match(r'(\d+)月(\d+)日', value)
            if match:
                month = int(match.group(1))
                day = int(match.group(2))
                # 创建完整日期
                full_date = datetime(year, month, day)
                return full_date.strftime('%Y-%m-%d')
        return value
    except (ValueError, TypeError):
        return value

def get_column_letter(index):
    """Convert column index to Excel column letter"""
    result = ""
    while index > 0:
        index -= 1
        result = chr(index % 26 + ord('A')) + result
        index //= 26
    return result

def get_column_index(letter):
    """Convert Excel column letter to index"""
    result = 0
    for char in letter:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result

def forward_fill_row(row):
    """Forward fill empty values in a row"""
    last_value = None
    filled_row = []
    for value in row:
        if pd.isna(value) or value == '':
            filled_row.append(last_value)
        else:
            last_value = value
            filled_row.append(value)
    return filled_row

def remove_trailing_empty_rows(df):
    """Remove trailing rows where all values are empty"""
    # 检查每一行是否所有值都为空
    empty_rows = df.apply(lambda row: row.isna().all() or (row == '').all(), axis=1)
    
    # 找到最后一个非空行的索引
    last_non_empty = len(df) - 1
    while last_non_empty >= 0 and empty_rows.iloc[last_non_empty]:
        last_non_empty -= 1
    
    # 如果找到非空行，保留到该行
    if last_non_empty >= 0:
        return df.iloc[:last_non_empty + 1]
    return df

# 读取 Excel 文件，使用data_only=True来获取公式计算后的值
file_path = "./uploads/raw_data.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

# 从工作表名称获取年份
sheet_name = sheet.title
year = int(sheet_name[:-1])  # 假设工作表名称就是年份
print(f"Sheet name: {sheet_name}, Year: {year}")

# 定义列范围
column_ranges = []
start_col = 'A'
end_col = 'DP'
current_col = start_col

while get_column_index(current_col) <= get_column_index(end_col):
    start_idx = get_column_index(current_col)
    end_idx = min(start_idx + 9, get_column_index(end_col))  # 每组10列
    end_letter = get_column_letter(end_idx)
    column_ranges.append((current_col, end_letter))
    current_col = get_column_letter(end_idx + 1)

# 为每个列范围创建CSV文件
for i, (start_col, end_col) in enumerate(column_ranges):
    print(f"Processing columns {start_col} to {end_col}")
    
    # 获取列范围的数据
    data = []
    for row in sheet.iter_rows(min_col=get_column_index(start_col), 
                             max_col=get_column_index(end_col),
                             values_only=True):
        processed_row = []
        for cell_value in row:
            # 如果是第一列且是日期，处理日期格式
            if len(processed_row) == 0 and cell_value is not None:
                cell_value = process_date(cell_value, year)
            processed_row.append(cell_value)
        data.append(processed_row)
    
    # 转换为DataFrame
    df = pd.DataFrame(data)
    
    # 获取第三行的机场名作为文件名
    airport_name = df.iloc[2, 0]  # 第三行第一列的值
    if pd.isna(airport_name) or not isinstance(airport_name, str):
        airport_name = f"part_{i+1}"
    else:
        # 清理文件名，移除不合法字符
        airport_name = re.sub(r'[<>:"/\\|?*]', '', airport_name)
        airport_name = airport_name.strip()
    
    # 处理第四行（索引为3）的空值
    df.iloc[3] = forward_fill_row(df.iloc[3])
    
    # 删除前两行
    df = df.iloc[2:]
    
    # 删除尾部的空行
    df = remove_trailing_empty_rows(df)
    
    # 导出为CSV
    output_filename = f"./uploads/{airport_name}.csv"
    df.to_csv(output_filename, index=False, header=False)
    print(f"Created {output_filename}")

print("所有文件处理完成！")